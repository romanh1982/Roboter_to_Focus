#!/usr/bin/env python3
"""
KUKA Robot Movement Profile Viewer with Laser Tracker Distance Calculator
=========================================================================
Loads KUKA KRL .src/.dat files, visualizes the robot TCP path,
computes the Euclidean distance from a configurable Laser Tracker position
to each waypoint, and produces a distance-over-time profile suitable
for feeding into the MiniVario focus test generator.

Exports:
  - CSV with full data
  - CSV with MiniVario distance profile (time; distance)
  - C++ header file with const lookup table for the Profile-Replay test mode

MiniVario distance range: 1500 mm .. 25000 mm  (MIN_INPUT_DIST .. MAX_INPUT_DIST)
Distance generator rate:  100 Hz  (10 ms per cycle)
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import re
import math
import csv
import os
from datetime import datetime

import numpy as np
import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
import matplotlib.pyplot as plt

# Tighter default margins so every plot auto-fits its data snugly
matplotlib.rcParams["axes.xmargin"] = 0.01
matplotlib.rcParams["axes.ymargin"] = 0.03
from commutator_animation_tab import CommutatorAnimationTab


# ---------------------------------------------------------------------------
#  Constants matching VarioZoomDefs.h
# ---------------------------------------------------------------------------
MV_MIN_DIST        = 1500    # mm  (MIN_INPUT_DIST)
MV_MAX_DIST        = 25000   # mm  (MAX_INPUT_DIST)
GENERATOR_RATE_HZ  = 100     # distance generator cycle rate
CYCLE_TIME_MS      = 10      # 1000 / GENERATOR_RATE_HZ
MAX_DELTA_DIST     = 60      # mm per 100 Hz cycle (MAX_DELTA_DIST_PER_CYCLE_POS_100)

# ---------------------------------------------------------------------------
#  Kinematic chain constants (from commutation granularity analysis)
#    Pulley diameter:  9 mm  →  circumference = π × 9 = 28.2743 mm
#    Gear ratio:       161 : 1
#    Motor rev/mm:     161 / (π × 9) = 5.6942
# ---------------------------------------------------------------------------
PULLEY_DIAMETER_MM  = 9.0
GEAR_RATIO          = 161.0
MOTOR_REV_PER_MM    = GEAR_RATIO / (math.pi * PULLEY_DIAMETER_MM)  # 5.6942 rev/mm
COMMUTATOR_SEGMENTS = 5       # 5-segment commutator
SECTOR_ANGLE_DEG    = 360.0 / COMMUTATOR_SEGMENTS  # 72°
LINEAR_PER_SECTOR_UM = 1000.0 / (MOTOR_REV_PER_MM * COMMUTATOR_SEGMENTS)  # ~35.1 µm
INCR_TO_MM          = 0.0005  # 1 increment = 0.5 µm = 0.0005 mm
ZOOM_TABLE_DIST_MIN = 1500   # mm – table index 0
ZOOM_TABLE_DIST_MAX = 25000  # mm – last table index
ZOOM_TABLE_ENTRIES  = 2351   # actual array size (indices 0..2350)
ZOOM_TABLE_DIST_STEP = 10.0  # mm per table index step

# Default path to zoom table header – try several locations
def _find_default_zoom_table():
    """Search for varioZoomTable25.h in likely locations."""
    candidates = [
        # Same directory as this script
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "varioZoomTable25.h"),
        # Current working directory
        os.path.join(os.getcwd(), "varioZoomTable25.h"),
        # Original PyCharm project path
        os.path.join(
            os.path.expanduser("~"),
            "PycharmProjects", "CommAnalyser", "cpp_files", "genie",
            "sensor", "fu", "vario", "public", "varioZoomTable25.h"
        ),
    ]
    for path in candidates:
        if os.path.isfile(path):
            return path
    return candidates[0]  # return first candidate as fallback (will fail gracefully)

DEFAULT_ZOOM_TABLE_PATH = _find_default_zoom_table()


def clamp_mv(dist_mm):
    return max(MV_MIN_DIST, min(MV_MAX_DIST, dist_mm))


# ---------------------------------------------------------------------------
#  Zoom table parsing & motor profile computation
# ---------------------------------------------------------------------------
def parse_zoom_tables(header_path):
    """Parse zoomTableS8[] and zoomTableS23[] from the C++ header file."""
    with open(header_path, "r", errors="replace") as f:
        text = f.read()

    def _extract_array(name):
        pattern = rf"const\s+long\s+{name}\s*\[\s*\]\s*=\s*\{{(.*?)\}}"
        m = re.search(pattern, text, re.DOTALL)
        if not m:
            raise ValueError(f"Could not find array '{name}' in {header_path}")
        body = m.group(1)
        # Strip comment lines (// ...) to avoid picking up stray numbers
        body = re.sub(r"//[^\n]*", "", body)
        nums = re.findall(r"-?\d+", body)
        return [int(n) for n in nums]

    return _extract_array("zoomTableS8"), _extract_array("zoomTableS23")


def dist_to_lens_pos(dist_mm, zoom_table):
    """Look up lens position (increments) for a given focus distance using the zoom table.
    Interpolates between table entries for sub-index distances."""
    idx_f = (dist_mm - ZOOM_TABLE_DIST_MIN) / ZOOM_TABLE_DIST_STEP
    idx_f = max(0.0, min(idx_f, len(zoom_table) - 1))
    idx_lo = int(idx_f)
    idx_hi = min(idx_lo + 1, len(zoom_table) - 1)
    frac = idx_f - idx_lo
    return zoom_table[idx_lo] + frac * (zoom_table[idx_hi] - zoom_table[idx_lo])


def lens_pos_to_linear_mm(increments):
    """Convert lens position in increments to linear displacement in mm."""
    return increments * INCR_TO_MM


def linear_mm_to_motor_angle(linear_mm):
    """Convert linear carriage displacement (mm) to motor angle (degrees)."""
    return linear_mm * MOTOR_REV_PER_MM * 360.0


def compute_motor_profiles(lut, zoom_s8, zoom_s23):
    """Compute M1/M2 lens positions and motor angles for each LUT distance entry.

    Returns a dict with arrays for each motor:
      m1_incr, m2_incr    – lens positions in increments
      m1_mm, m2_mm        – linear displacements in mm
      m1_deg, m2_deg      – motor angles in degrees
      m1_sectors, m2_sectors – cumulative commutation sectors
    """
    m1_incr, m2_incr = [], []
    m1_mm, m2_mm = [], []
    m1_deg, m2_deg = [], []

    for dist in lut:
        # Lens positions from zoom tables
        p1 = dist_to_lens_pos(dist, zoom_s8)
        p2 = dist_to_lens_pos(dist, zoom_s23)
        m1_incr.append(p1)
        m2_incr.append(p2)

        # Linear displacement
        l1 = lens_pos_to_linear_mm(p1)
        l2 = lens_pos_to_linear_mm(p2)
        m1_mm.append(l1)
        m2_mm.append(l2)

        # Motor angle
        m1_deg.append(linear_mm_to_motor_angle(l1))
        m2_deg.append(linear_mm_to_motor_angle(l2))

    # Commutation sectors (angle / 72°)
    m1_sectors = [a / 72.0 for a in m1_deg]
    m2_sectors = [a / 72.0 for a in m2_deg]

    return dict(
        m1_incr=m1_incr, m2_incr=m2_incr,
        m1_mm=m1_mm, m2_mm=m2_mm,
        m1_deg=m1_deg, m2_deg=m2_deg,
        m1_sectors=m1_sectors, m2_sectors=m2_sectors,
    )


# ---------------------------------------------------------------------------
#  Parser
# ---------------------------------------------------------------------------
def parse_kuka_files(src_path, dat_path):
    with open(dat_path, "r", errors="replace") as f:
        dat_text = f.read()
    with open(src_path, "r", errors="replace") as f:
        src_text = f.read()

    positions = {}
    for m in re.finditer(
        r"DECL\s+E6POS\s+(XP\d+)\s*=\s*\{"
        r"X\s+([0-9.\-]+),\s*Y\s+([0-9.\-]+),\s*Z\s+([0-9.\-]+),\s*"
        r"A\s+([0-9.\-]+),\s*B\s+([0-9.\-]+),\s*C\s+([0-9.\-]+)",
        dat_text,
    ):
        positions[m.group(1)] = dict(
            X=float(m.group(2)), Y=float(m.group(3)), Z=float(m.group(4)),
            A=float(m.group(5)), B=float(m.group(6)), C=float(m.group(7)),
        )

    moves = re.findall(
        r";FOLD\s+(SPTP|SLIN|SCIRC)\s+(P\d+)(\s+CONT)?\s+Vel=([0-9.]+)\s+(% |m/s)",
        src_text,
    )

    waypoints = []
    for i, (cmd, pname, cont, vel, unit) in enumerate(moves):
        xp = "X" + pname
        pos = positions.get(xp, {})
        if not pos:
            continue
        waypoints.append(dict(
            step=i + 1, point=pname, cmd=cmd,
            vel_value=float(vel),
            vel_unit="%" if "%" in unit else "m/s",
            blending="CONT" if cont.strip() else "STOP",
            X=pos["X"], Y=pos["Y"], Z=pos["Z"],
            A=pos["A"], B=pos["B"], C=pos["C"],
        ))
    return waypoints


# ---------------------------------------------------------------------------
#  Distance & time estimation
# ---------------------------------------------------------------------------
def compute_distances_and_time(waypoints, tracker_pos):
    tx, ty, tz = tracker_pos
    results = []
    cum_time = 0.0
    for i, wp in enumerate(waypoints):
        d_tracker = math.sqrt(
            (wp["X"] - tx) ** 2 + (wp["Y"] - ty) ** 2 + (wp["Z"] - tz) ** 2
        )
        if i == 0:
            seg_dist = 0.0
            seg_time = 0.0
        else:
            prev = waypoints[i - 1]
            seg_dist = math.sqrt(
                (wp["X"] - prev["X"]) ** 2 +
                (wp["Y"] - prev["Y"]) ** 2 +
                (wp["Z"] - prev["Z"]) ** 2
            )
            if wp["vel_unit"] == "m/s":
                v_mm_s = wp["vel_value"] * 1000.0
            else:
                v_mm_s = (wp["vel_value"] / 100.0) * 2000.0
            seg_time = seg_dist / max(v_mm_s, 0.001)
        cum_time += seg_time
        results.append(dict(
            step=wp["step"], point=wp["point"], cmd=wp["cmd"],
            vel=f'{wp["vel_value"]:.3f} {wp["vel_unit"]}',
            vel_value=wp["vel_value"], vel_unit=wp["vel_unit"],
            blending=wp["blending"],
            X=wp["X"], Y=wp["Y"], Z=wp["Z"],
            dist_to_tracker=d_tracker, segment_dist=seg_dist, time=cum_time,
        ))
    return results


# ---------------------------------------------------------------------------
#  C++ LUT generation  –  resample to 100 Hz cycles
# ---------------------------------------------------------------------------
def resample_to_100hz(results):
    if not results:
        return []
    times = [r["time"] for r in results]
    dists = [r["dist_to_tracker"] for r in results]
    total_time = times[-1]
    n_cycles = int(math.ceil(total_time * GENERATOR_RATE_HZ)) + 1
    dt = 1.0 / GENERATOR_RATE_HZ

    lut = []
    j = 0
    for ci in range(n_cycles):
        t = ci * dt
        while j < len(times) - 2 and times[j + 1] < t:
            j += 1
        if j >= len(times) - 1:
            d = dists[-1]
        else:
            t0, t1 = times[j], times[j + 1]
            d0, d1 = dists[j], dists[j + 1]
            frac = (t - t0) / (t1 - t0) if abs(t1 - t0) > 1e-9 else 0
            d = d0 + frac * (d1 - d0)
        lut.append(int(round(clamp_mv(d))))

    # Rate-limit: max step per cycle
    for i in range(1, len(lut)):
        delta = lut[i] - lut[i - 1]
        if delta > MAX_DELTA_DIST:
            lut[i] = lut[i - 1] + MAX_DELTA_DIST
        elif delta < -MAX_DELTA_DIST:
            lut[i] = lut[i - 1] - MAX_DELTA_DIST
        lut[i] = clamp_mv(lut[i])
    return lut


def generate_cpp_header(lut, src_name="", tracker_pos=(0, 0, 0)):
    n = len(lut)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    guard = "VARIO_PROFILE_REPLAY_LUT_H_"

    lines = []
    lines.append(f"// Auto-generated by kuka_movement_viewer.py  ({now})")
    lines.append(f"// Source program: {src_name}")
    lines.append(f"// Tracker position: X={tracker_pos[0]:.1f}  Y={tracker_pos[1]:.1f}  Z={tracker_pos[2]:.1f} mm")
    lines.append(f"// Generator rate: {GENERATOR_RATE_HZ} Hz  ({CYCLE_TIME_MS} ms / cycle)")
    lines.append(f"// Distance range: {min(lut)} .. {max(lut)} mm")
    lines.append(f"// Duration: {n} cycles = {n * CYCLE_TIME_MS / 1000.0:.2f} s")
    lines.append(f"// Rate-limited to max {MAX_DELTA_DIST} mm/cycle")
    lines.append("")
    lines.append(f"#ifndef {guard}")
    lines.append(f"#define {guard}")
    lines.append("")
    lines.append(f"#define PROFILE_REPLAY_LUT_SIZE   {n}")
    lines.append(f"#define PROFILE_REPLAY_CYCLE_MS   {CYCLE_TIME_MS}")
    lines.append("")
    lines.append("// Distance values in mm, one entry per 100 Hz generator cycle.")
    lines.append("// Usage: testDistanceRT = profileReplayLUT[profileReplayIndex];")
    lines.append(f"const long profileReplayLUT[PROFILE_REPLAY_LUT_SIZE] = {{")

    for row_start in range(0, n, 10):
        row_end = min(row_start + 10, n)
        vals = ", ".join(f"{lut[k]:5d}" for k in range(row_start, row_end))
        comma = "," if row_end < n else ""
        lines.append(f"  {vals}{comma}")

    lines.append("};")
    lines.append("")
    lines.append(f"#endif // {guard}")
    lines.append("")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
#  Main Application
# ---------------------------------------------------------------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("KUKA Movement Profile \u2013 Laser Tracker Distance Viewer")
        self.geometry("1500x920")
        self.minsize(1200, 700)
        self.waypoints = []
        self.results = []
        self.lut = []
        self.motor = {}          # motor profile data from compute_motor_profiles
        self.zoom_s8 = None      # parsed zoom tables
        self.zoom_s23 = None
        self.src_name = ""
        self._load_zoom_tables(DEFAULT_ZOOM_TABLE_PATH)
        self._build_ui()

    def _load_zoom_tables(self, path):
        """Load zoom tables from C++ header. Silently skip if file not found."""
        try:
            self.zoom_s8, self.zoom_s23 = parse_zoom_tables(path)
        except Exception:
            self.zoom_s8 = None
            self.zoom_s23 = None

    def _build_ui(self):
        ctrl = ttk.Frame(self, padding=6)
        ctrl.pack(fill=tk.X)
        ttk.Button(ctrl, text="Load .src / .dat files \u2026", command=self._load_files).pack(side=tk.LEFT, padx=4)
        ttk.Separator(ctrl, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)
        ttk.Label(ctrl, text="Tracker pos (X Y Z) [mm]:").pack(side=tk.LEFT)
        self.tracker_x = tk.DoubleVar(value=3800.0)
        self.tracker_y = tk.DoubleVar(value=-1000.0)
        self.tracker_z = tk.DoubleVar(value=1500.0)
        for var, label in [(self.tracker_x, "X"), (self.tracker_y, "Y"), (self.tracker_z, "Z")]:
            ttk.Label(ctrl, text=f" {label}:").pack(side=tk.LEFT)
            ttk.Entry(ctrl, textvariable=var, width=9).pack(side=tk.LEFT, padx=2)
        ttk.Button(ctrl, text="Recalculate", command=self._recalculate).pack(side=tk.LEFT, padx=8)

        exp = ttk.Frame(self, padding=(6, 0, 6, 4))
        exp.pack(fill=tk.X)
        ttk.Label(exp, text="Export:", foreground="gray").pack(side=tk.LEFT)
        ttk.Button(exp, text="CSV (full data)", command=self._export_csv).pack(side=tk.LEFT, padx=4)
        ttk.Button(exp, text="CSV (MV distance profile)", command=self._export_mv_profile).pack(side=tk.LEFT, padx=4)
        ttk.Button(exp, text="C++ Header (Profile Replay LUT)", command=self._export_cpp_lut).pack(side=tk.LEFT, padx=4)
        ttk.Button(exp, text="CSV (Motor Profiles)", command=self._export_motor_csv).pack(side=tk.LEFT, padx=4)
        ttk.Button(exp, text="Excel (Motor Profile for Supplier)", command=self._export_motor_xlsx).pack(side=tk.LEFT, padx=4)
        ttk.Separator(exp, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)
        zt_status = "loaded" if self.zoom_s8 else "NOT FOUND"
        self.zoom_status_var = tk.StringVar(value=f"Zoom tables: {zt_status}")
        ttk.Label(exp, textvariable=self.zoom_status_var, foreground="gray").pack(side=tk.LEFT, padx=4)
        ttk.Button(exp, text="Load Zoom Table\u2026", command=self._browse_zoom_table).pack(side=tk.LEFT, padx=2)

        self.status_var = tk.StringVar(value="No files loaded.")
        ttk.Label(self, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W,
                  padding=4).pack(fill=tk.X, side=tk.BOTTOM)

        self.nb = ttk.Notebook(self)
        self.nb.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        self.tab_path = ttk.Frame(self.nb); self.nb.add(self.tab_path, text="  Robot Path  ")
        self.tab_dist = ttk.Frame(self.nb); self.nb.add(self.tab_dist, text="  Tracker Distance vs Time  ")
        self.tab_mv   = ttk.Frame(self.nb); self.nb.add(self.tab_mv,   text="  MiniVario Focus Profile  ")
        self.tab_lut  = ttk.Frame(self.nb); self.nb.add(self.tab_lut,  text="  100 Hz LUT Preview  ")
        self.tab_lens = ttk.Frame(self.nb); self.nb.add(self.tab_lens, text="  M1/M2 Lens Profiles  ")
        self.tab_motor= ttk.Frame(self.nb); self.nb.add(self.tab_motor,text="  Motor Angle Profiles  ")
        self.tab_table= ttk.Frame(self.nb); self.nb.add(self.tab_table,text="  Data Table  ")
        self.tab_comm = ttk.Frame(self.nb)
        self.nb.add(self.tab_comm, text="  Commutator Animation  ")
        self.comm_anim = CommutatorAnimationTab(self.tab_comm, self)

        self._init_path_tab()
        self._init_dist_tab()
        self._init_mv_tab()
        self._init_lut_tab()
        self._init_lens_tab()
        self._init_motor_tab()
        self._init_table_tab()

    # --- tabs ---------------------------------------------------------------
    def _init_path_tab(self):
        tb = ttk.Frame(self.tab_path); tb.pack(fill=tk.X)
        self.path_view_var = tk.StringVar(value="xy")
        for v, l in [("xy","Top (X-Y)"),("xz","Side (X-Z)"),("yz","Front (Y-Z)")]:
            ttk.Radiobutton(tb, text=l, variable=self.path_view_var, value=v,
                            command=self._draw_path).pack(side=tk.LEFT, padx=6, pady=4)
        self.path_color_var = tk.StringVar(value="type")
        ttk.Label(tb, text="   Color:").pack(side=tk.LEFT)
        for v, l in [("type","Move type"),("vel","Velocity"),("blend","Blending")]:
            ttk.Radiobutton(tb, text=l, variable=self.path_color_var, value=v,
                            command=self._draw_path).pack(side=tk.LEFT, padx=4)
        self.show_tracker_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(tb, text="Show tracker", variable=self.show_tracker_var,
                        command=self._draw_path).pack(side=tk.LEFT, padx=10)
        self.fig_path = Figure(figsize=(12,6), dpi=100)
        self.ax_path = self.fig_path.add_subplot(111)
        self.canvas_path = FigureCanvasTkAgg(self.fig_path, master=self.tab_path)
        NavigationToolbar2Tk(self.canvas_path, self.tab_path)
        self.canvas_path.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        self._dragging_tracker = False
        self.canvas_path.mpl_connect("button_press_event", self._on_path_press)
        self.canvas_path.mpl_connect("motion_notify_event", self._on_path_motion)
        self.canvas_path.mpl_connect("button_release_event", self._on_path_release)

    def _init_dist_tab(self):
        self.fig_dist = Figure(figsize=(12,5), dpi=100)
        self.ax_dist = self.fig_dist.add_subplot(111)
        self.canvas_dist = FigureCanvasTkAgg(self.fig_dist, master=self.tab_dist)
        NavigationToolbar2Tk(self.canvas_dist, self.tab_dist)
        self.canvas_dist.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def _init_mv_tab(self):
        ttk.Label(self.tab_mv, text="MiniVario distance range: 1500 \u2013 25000 mm.  "
                  "Values outside this range are clamped.", foreground="gray",
                  padding=4).pack(anchor=tk.W)
        self.fig_mv = Figure(figsize=(12,5), dpi=100)
        self.ax_mv_dist = self.fig_mv.add_subplot(211)
        self.ax_mv_speed = self.fig_mv.add_subplot(212, sharex=self.ax_mv_dist)
        self.canvas_mv = FigureCanvasTkAgg(self.fig_mv, master=self.tab_mv)
        NavigationToolbar2Tk(self.canvas_mv, self.tab_mv)
        self.canvas_mv.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def _init_lut_tab(self):
        info = ttk.Frame(self.tab_lut, padding=4); info.pack(fill=tk.X)
        ttk.Label(info, text="Preview of the resampled 100 Hz lookup table for the "
                  "Profile-Replay test mode.  Rate-limited to \u00b160 mm/cycle.",
                  foreground="gray").pack(anchor=tk.W)
        self.lut_info_var = tk.StringVar(value="")
        ttk.Label(info, textvariable=self.lut_info_var, foreground="#534ab7").pack(anchor=tk.W)
        self.fig_lut = Figure(figsize=(12,5), dpi=100)
        self.ax_lut_dist = self.fig_lut.add_subplot(211)
        self.ax_lut_delta = self.fig_lut.add_subplot(212, sharex=self.ax_lut_dist)
        self.canvas_lut = FigureCanvasTkAgg(self.fig_lut, master=self.tab_lut)
        NavigationToolbar2Tk(self.canvas_lut, self.tab_lut)
        self.canvas_lut.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def _init_lens_tab(self):
        ttk.Label(self.tab_lens, text="M1 (ZOOM1/S8) and M2 (ZOOM2/S23) lens carriage positions "
                  "derived from the zoom lookup tables.  Units: increments (0.5 \u00b5m each).",
                  foreground="gray", padding=4).pack(anchor=tk.W)
        self.lens_info_var = tk.StringVar(value="")
        ttk.Label(self.tab_lens, textvariable=self.lens_info_var, foreground="#534ab7",
                  padding=(4,0)).pack(anchor=tk.W)
        self.fig_lens = Figure(figsize=(12,5), dpi=100)
        self.ax_lens_pos = self.fig_lens.add_subplot(211)
        self.ax_lens_mm  = self.fig_lens.add_subplot(212, sharex=self.ax_lens_pos)
        self.canvas_lens = FigureCanvasTkAgg(self.fig_lens, master=self.tab_lens)
        NavigationToolbar2Tk(self.canvas_lens, self.tab_lens)
        self.canvas_lens.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def _init_motor_tab(self):
        ttk.Label(self.tab_motor, text="Motor angular position derived from lens travel via "
                  f"kinematic chain (d={PULLEY_DIAMETER_MM:.0f} mm, ratio={GEAR_RATIO:.0f}:1 "
                  f"\u2192 {MOTOR_REV_PER_MM:.4f} rev/mm).  "
                  f"Commutation sector = {SECTOR_ANGLE_DEG:.0f}\u00b0 "
                  f"({COMMUTATOR_SEGMENTS}-segment) \u2248 {LINEAR_PER_SECTOR_UM:.1f} \u00b5m linear.",
                  foreground="gray", padding=4).pack(anchor=tk.W)
        self.motor_info_var = tk.StringVar(value="")
        ttk.Label(self.tab_motor, textvariable=self.motor_info_var, foreground="#534ab7",
                  padding=(4,0)).pack(anchor=tk.W)
        self.fig_motor = Figure(figsize=(12,5), dpi=100)
        self.ax_motor_deg = self.fig_motor.add_subplot(211)
        self.ax_motor_sec = self.fig_motor.add_subplot(212, sharex=self.ax_motor_deg)
        self.canvas_motor = FigureCanvasTkAgg(self.fig_motor, master=self.tab_motor)
        NavigationToolbar2Tk(self.canvas_motor, self.tab_motor)
        self.canvas_motor.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def _init_table_tab(self):
        cols = ("step","point","cmd","vel","blend","X","Y","Z","dist_tracker","seg_dist","time")
        self.tree = ttk.Treeview(self.tab_table, columns=cols, show="headings", height=25)
        hdr = {"step":"Step","point":"Point","cmd":"Cmd","vel":"Velocity","blend":"Blend",
               "X":"X [mm]","Y":"Y [mm]","Z":"Z [mm]",
               "dist_tracker":"Dist Tracker [mm]","seg_dist":"Seg Dist [mm]","time":"Time [s]"}
        wid = {"step":50,"point":55,"cmd":50,"vel":100,"blend":55,
               "X":85,"Y":85,"Z":85,"dist_tracker":120,"seg_dist":100,"time":80}
        for c in cols:
            self.tree.heading(c, text=hdr[c])
            self.tree.column(c, width=wid[c], anchor=tk.E if c not in ("point","cmd","blend","vel") else tk.CENTER)
        vsb = ttk.Scrollbar(self.tab_table, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.pack(fill=tk.BOTH, expand=True)

    # --- actions ------------------------------------------------------------
    def _load_files(self):
        src_path = filedialog.askopenfilename(title="Select KUKA .src file",
            filetypes=[("KUKA Source","*.src"),("All","*.*")])
        if not src_path: return
        base = os.path.splitext(src_path)[0]
        dat_path = base + ".dat"
        if not os.path.isfile(dat_path):
            dat_path = filedialog.askopenfilename(title="Select matching .dat file",
                filetypes=[("KUKA Data","*.dat"),("All","*.*")])
        if not dat_path or not os.path.isfile(dat_path):
            messagebox.showerror("Error", "No .dat file selected or found."); return
        try:
            self.waypoints = parse_kuka_files(src_path, dat_path)
        except Exception as e:
            messagebox.showerror("Parse Error", str(e)); return
        if not self.waypoints:
            messagebox.showwarning("Warning", "No waypoints parsed."); return
        self.src_name = os.path.basename(src_path)
        self.status_var.set(f"Loaded {len(self.waypoints)} waypoints from {self.src_name}")
        self._recalculate()

    def _browse_zoom_table(self):
        path = filedialog.askopenfilename(title="Select varioZoomTable25.h",
            filetypes=[("C/C++ Header","*.h"),("All","*.*")])
        if not path:
            return
        self._load_zoom_tables(path)
        if self.zoom_s8:
            self.zoom_status_var.set(f"Zoom tables: loaded ({len(self.zoom_s8)} entries)")
            if self.lut:
                self._recalculate()
        else:
            self.zoom_status_var.set("Zoom tables: PARSE ERROR")
            messagebox.showerror("Error", "Could not parse zoom tables from selected file.")

    def _recalculate(self):
        if not self.waypoints: return
        tp = (self.tracker_x.get(), self.tracker_y.get(), self.tracker_z.get())
        self.results = compute_distances_and_time(self.waypoints, tp)
        self.lut = resample_to_100hz(self.results)
        # Compute motor profiles if zoom tables are loaded
        if self.zoom_s8 and self.zoom_s23 and self.lut:
            self.motor = compute_motor_profiles(self.lut, self.zoom_s8, self.zoom_s23)
        else:
            self.motor = {}
        self._draw_path(); self._draw_dist(); self._draw_mv()
        self._draw_lut(); self._draw_lens(); self._draw_motor()
        self._fill_table()
        if hasattr(self, 'comm_anim'):
            self.comm_anim.update_data()
        dists = [r["dist_to_tracker"] for r in self.results]
        zt = "  |  Motor profiles: OK" if self.motor else "  |  Motor profiles: no zoom tables"
        self.status_var.set(
            f"{len(self.results)} waypoints  |  Tracker dist: {min(dists):.0f}\u2013{max(dists):.0f} mm  |  "
            f"Time \u2248 {self.results[-1]['time']:.1f} s  |  LUT: {len(self.lut)} entries @ 100 Hz{zt}")

    # --- drawing ------------------------------------------------------------
    def _get_color(self, wp, mode):
        if mode == "type":
            return "#2166ac" if wp["cmd"] == "SPTP" else "#1b9e77"
        elif mode == "blend":
            return "#e24b4a" if wp["blending"] == "STOP" else "#2166ac"
        else:
            v = wp.get("vel_value", 0); u = wp.get("vel_unit", "%")
            if u == "%": return "#1b9e77" if v >= 90 else ("#ef9f27" if v >= 60 else "#e24b4a")
            else:        return "#1b9e77" if v >= 0.2 else ("#ef9f27" if v >= 0.1 else "#e24b4a")

    def _draw_path(self):
        if not self.waypoints: return
        ax = self.ax_path; ax.clear()
        view = self.path_view_var.get(); cm = self.path_color_var.get()
        hk, vk = {"xy":("X","Y"),"xz":("X","Z"),"yz":("Y","Z")}[view]
        wps = self.waypoints
        hs = [w[hk] for w in wps]; vs = [w[vk] for w in wps]
        for i in range(len(wps)-1):
            ax.plot([hs[i],hs[i+1]],[vs[i],vs[i+1]], color=self._get_color(wps[i+1],cm), lw=1.2, alpha=0.6)
        for i, wp in enumerate(wps):
            ax.plot(wp[hk], wp[vk], "o", color=self._get_color(wp,cm), ms=5 if wp["blending"]=="STOP" else 2)
        ax.plot(hs[0],vs[0],"o",color="#1b9e77",ms=10,mfc="none",mew=2)
        ax.annotate("START",(hs[0],vs[0]),fontsize=8,fontweight="bold",color="#1b9e77",xytext=(8,8),textcoords="offset points")
        if self.show_tracker_var.get():
            tp=(self.tracker_x.get(),self.tracker_y.get(),self.tracker_z.get())
            th=tp[{"X":0,"Y":1,"Z":2}[hk]]; tv=tp[{"X":0,"Y":1,"Z":2}[vk]]
            ax.plot(th,tv,"s",color="#e24b4a",ms=12,mew=2,mfc="none")
            ax.annotate("TRACKER",(th,tv),fontsize=8,fontweight="bold",color="#e24b4a",xytext=(10,-12),textcoords="offset points")
            ax.plot([th,hs[0]],[tv,vs[0]],"--",color="#e24b4a",lw=0.8,alpha=0.5)
        ax.set_xlabel(f"{hk} [mm]"); ax.set_ylabel(f"{vk} [mm]")
        ax.set_title(f"Robot TCP Path \u2013 {view.upper()} view")
        ax.set_aspect("equal", adjustable="datalim"); ax.grid(True, alpha=0.3)
        self.fig_path.tight_layout(); self.canvas_path.draw()

    # --- tracker drag on robot path ----------------------------------------
    def _path_axes_keys(self):
        return {"xy":("X","Y"),"xz":("X","Z"),"yz":("Y","Z")}[self.path_view_var.get()]

    def _tracker_screen_pos(self):
        hk, vk = self._path_axes_keys()
        tp = (self.tracker_x.get(), self.tracker_y.get(), self.tracker_z.get())
        idx = {"X":0,"Y":1,"Z":2}
        return tp[idx[hk]], tp[idx[vk]]

    def _on_path_press(self, event):
        if event.inaxes is not self.ax_path or event.button != 1: return
        if not self.show_tracker_var.get(): return
        if event.xdata is None or event.ydata is None: return
        # hit test in pixel space (~15 px radius)
        th, tv = self._tracker_screen_pos()
        tx_pix, ty_pix = self.ax_path.transData.transform((th, tv))
        dx = tx_pix - event.x; dy = ty_pix - event.y
        if (dx*dx + dy*dy) ** 0.5 <= 15:
            self._dragging_tracker = True

    def _on_path_motion(self, event):
        if not self._dragging_tracker: return
        if event.inaxes is not self.ax_path: return
        if event.xdata is None or event.ydata is None: return
        hk, vk = self._path_axes_keys()
        var_map = {"X": self.tracker_x, "Y": self.tracker_y, "Z": self.tracker_z}
        var_map[hk].set(round(float(event.xdata), 1))
        var_map[vk].set(round(float(event.ydata), 1))
        self._draw_path()

    def _on_path_release(self, event):
        if not self._dragging_tracker: return
        self._dragging_tracker = False
        self._recalculate()

    def _draw_dist(self):
        if not self.results: return
        ax = self.ax_dist; ax.clear()
        ts=[r["time"] for r in self.results]; ds=[r["dist_to_tracker"] for r in self.results]
        ax.plot(ts, ds, "-", color="#2166ac", lw=1.2, label="Distance Tracker\u2192TCP")
        for r in self.results:
            if r["blending"]=="STOP": ax.axvline(r["time"],color="#e24b4a",lw=0.5,alpha=0.4)
        # Lock Y-limits to the actual trajectory data before adding MV-range
        # reference shading, otherwise the 1500-25000 mm span would expand the axis.
        d_lo, d_hi = min(ds), max(ds); d_pad = max((d_hi - d_lo) * 0.08, 10.0)
        ax.set_ylim(d_lo - d_pad, d_hi + d_pad)
        ax.axhspan(MV_MIN_DIST, MV_MAX_DIST, alpha=0.08, color="#1b9e77",
                   label=f"MV range ({MV_MIN_DIST}\u2013{MV_MAX_DIST} mm)")
        ax.axhline(MV_MIN_DIST,color="#1b9e77",lw=0.8,ls="--",alpha=0.5)
        ax.axhline(MV_MAX_DIST,color="#1b9e77",lw=0.8,ls="--",alpha=0.5)
        ax.set_ylim(d_lo - d_pad, d_hi + d_pad)
        ax.set_xlabel("Estimated time [s]"); ax.set_ylabel("Distance [mm]")
        ax.set_title("Laser Tracker Distance to Scanner vs Time")
        ax.legend(fontsize=8); ax.grid(True, alpha=0.3)
        self.fig_dist.tight_layout(); self.canvas_dist.draw()

    def _draw_mv(self):
        if not self.results: return
        ts=[r["time"] for r in self.results]; raw=[r["dist_to_tracker"] for r in self.results]
        cl=[clamp_mv(d) for d in raw]
        sp=[0.0]+[abs(cl[i]-cl[i-1])/(ts[i]-ts[i-1]) if ts[i]-ts[i-1]>0.001 else 0 for i in range(1,len(cl))]
        ax1=self.ax_mv_dist; ax1.clear()
        ax1.plot(ts,cl,"-",color="#534ab7",lw=1.2,label="MV input distance (clamped)")
        ax1.plot(ts,raw,"-",color="#b4b2a9",lw=0.6,alpha=0.5,label="Raw tracker distance")
        d_lo = min(min(cl), min(raw)); d_hi = max(max(cl), max(raw))
        d_pad = max((d_hi - d_lo) * 0.08, 10.0)
        ax1.set_ylim(d_lo - d_pad, d_hi + d_pad)
        ax1.axhline(MV_MIN_DIST,color="#1b9e77",lw=0.8,ls="--",alpha=0.5)
        ax1.axhline(MV_MAX_DIST,color="#1b9e77",lw=0.8,ls="--",alpha=0.5)
        ax1.set_ylim(d_lo - d_pad, d_hi + d_pad)
        ax1.set_ylabel("Distance [mm]"); ax1.set_title("MiniVario Focus Distance Profile")
        ax1.legend(fontsize=8); ax1.grid(True,alpha=0.3)
        ax2=self.ax_mv_speed; ax2.clear()
        ax2.plot(ts,sp,"-",color="#d85a30",lw=1.0,label="|\u0394dist/\u0394t|")
        ax2.set_xlabel("Time [s]"); ax2.set_ylabel("Rate [mm/s]")
        ax2.set_title("Distance Change Rate"); ax2.legend(fontsize=8); ax2.grid(True,alpha=0.3)
        self.fig_mv.tight_layout(); self.canvas_mv.draw()

    def _draw_lut(self):
        if not self.lut: return
        n=len(self.lut); dt=CYCLE_TIME_MS/1000.0; ts=[i*dt for i in range(n)]
        self.lut_info_var.set(
            f"LUT size: {n} entries  |  Duration: {n*dt:.2f} s  |  "
            f"Range: {min(self.lut)}\u2013{max(self.lut)} mm  |  Memory: {n*4} bytes (long)")
        ax1=self.ax_lut_dist; ax1.clear()
        ax1.plot(ts,self.lut,"-",color="#534ab7",lw=0.8,label="profileReplayLUT[]")
        l_lo, l_hi = min(self.lut), max(self.lut)
        l_pad = max((l_hi - l_lo) * 0.08, 10.0)
        ax1.set_ylim(l_lo - l_pad, l_hi + l_pad)
        ax1.axhline(MV_MIN_DIST,color="#1b9e77",lw=0.8,ls="--",alpha=0.5)
        ax1.axhline(MV_MAX_DIST,color="#1b9e77",lw=0.8,ls="--",alpha=0.5)
        ax1.set_ylim(l_lo - l_pad, l_hi + l_pad)
        ax1.set_ylabel("Distance [mm]"); ax1.set_title("Profile Replay LUT \u2013 Distance per 100 Hz cycle")
        ax1.legend(fontsize=8); ax1.grid(True,alpha=0.3)
        deltas=[0]+[self.lut[i]-self.lut[i-1] for i in range(1,n)]
        ax2=self.ax_lut_delta; ax2.clear()
        ax2.plot(ts,deltas,"-",color="#d85a30",lw=0.6,label="\u0394dist per cycle [mm]")
        ax2.axhline(MAX_DELTA_DIST,color="#e24b4a",lw=0.8,ls="--",alpha=0.5,label=f"\u00b1{MAX_DELTA_DIST} mm limit")
        ax2.axhline(-MAX_DELTA_DIST,color="#e24b4a",lw=0.8,ls="--",alpha=0.5)
        ax2.set_xlabel("Time [s]"); ax2.set_ylabel("\u0394dist/cycle [mm]")
        ax2.set_title("Distance change per cycle (rate-limited)"); ax2.legend(fontsize=8); ax2.grid(True,alpha=0.3)
        self.fig_lut.tight_layout(); self.canvas_lut.draw()

    def _draw_lens(self):
        if not self.motor: return
        n = len(self.motor["m1_incr"])
        dt = CYCLE_TIME_MS / 1000.0
        ts = [i * dt for i in range(n)]

        self.lens_info_var.set(
            f"M1 (S8): {self.motor['m1_incr'][0]:.0f}\u2013{max(self.motor['m1_incr']):.0f} incr  "
            f"({self.motor['m1_mm'][0]:.3f}\u2013{max(self.motor['m1_mm']):.3f} mm)  |  "
            f"M2 (S23): {self.motor['m2_incr'][0]:.0f}\u2013{max(self.motor['m2_incr']):.0f} incr  "
            f"({self.motor['m2_mm'][0]:.3f}\u2013{max(self.motor['m2_mm']):.3f} mm)")

        ax1 = self.ax_lens_pos; ax1.clear()
        ax1.plot(ts, self.motor["m1_incr"], "-", color="#2166ac", lw=0.8, label="M1 / ZOOM1 (S8)")
        ax1.plot(ts, self.motor["m2_incr"], "-", color="#d85a30", lw=0.8, label="M2 / ZOOM2 (S23)")
        ax1.set_ylabel("Lens position [increments]")
        ax1.set_title("M1 / M2 Lens Carriage Positions (from Zoom Lookup Tables)")
        ax1.legend(fontsize=8); ax1.grid(True, alpha=0.3)

        ax2 = self.ax_lens_mm; ax2.clear()
        ax2.plot(ts, self.motor["m1_mm"], "-", color="#2166ac", lw=0.8, label="M1 linear [mm]")
        ax2.plot(ts, self.motor["m2_mm"], "-", color="#d85a30", lw=0.8, label="M2 linear [mm]")
        ax2.set_xlabel("Time [s]"); ax2.set_ylabel("Linear displacement [mm]")
        ax2.set_title("Lens Carriage Linear Displacement (1 incr = 0.5 \u00b5m)")
        ax2.legend(fontsize=8); ax2.grid(True, alpha=0.3)
        self.fig_lens.tight_layout(); self.canvas_lens.draw()

    def _apply_motor_yticks(self, ax, y_lo, y_hi):
        """Set major/minor Y ticks on the motor angle axis, auto-choosing label density.
        Major (labeled): every N×360° so that ~6-10 labels fit.
        Minor (unlabeled): every 360° (one revolution).
        When zoomed in below ~5 rev, minor switches to 72° (one commutation sector)."""
        y_range = y_hi - y_lo
        if y_range <= 0:
            return
        revs_visible = y_range / 360.0
        max_labels = 8

        # Fine grid adapts: 72° sectors when zoomed in, 360° revolutions when zoomed out
        if revs_visible <= 5:
            fine_step = 72.0
        else:
            fine_step = 360.0

        # Label step: multiples of 360° so labels don't overlap
        label_rev_step = max(1, int(np.ceil(revs_visible / max_labels)))
        label_step = label_rev_step * 360.0

        # Align ticks to multiples starting from 0
        label_lo = int(np.floor(y_lo / label_step)) * label_step
        label_ticks = np.arange(label_lo, y_hi + label_step, label_step)
        fine_lo = int(np.floor(y_lo / fine_step)) * fine_step
        fine_ticks = np.arange(fine_lo, y_hi + fine_step, fine_step)

        ax.set_yticks(label_ticks)
        ax.set_yticks(fine_ticks, minor=True)
        ax.set_yticklabels([f"{int(v)}\u00b0 ({int(v/360)} rev)" for v in label_ticks], fontsize=8)
        ax.set_ylim(y_lo, y_hi)
        ax.grid(True, which="major", alpha=0.6, lw=0.9)
        ax.grid(True, which="minor", alpha=0.15, lw=0.4)

    def _on_motor_xlim_changed(self, ax):
        """Callback: when shared X-axis changes, auto-rescale Y on both subplots."""
        if getattr(self, '_motor_updating', False):
            return
        self._motor_updating = True
        try:
            x_lo, x_hi = ax.get_xlim()
            mask = (self._motor_ts >= x_lo) & (self._motor_ts <= x_hi)
            if not np.any(mask):
                return

            # Rescale top subplot Y (motor angle)
            ax1 = self.ax_motor_deg
            vis_m1 = self._motor_m1_deg[mask]
            vis_m2 = self._motor_m2_deg[mask]
            y_lo = min(vis_m1.min(), vis_m2.min())
            y_hi = max(vis_m1.max(), vis_m2.max())
            margin = (y_hi - y_lo) * 0.05 if y_hi > y_lo else 180
            self._apply_motor_yticks(ax1, y_lo - margin, y_hi + margin)
            # Sync top right axis
            ylim = ax1.get_ylim()
            self._motor_ax1_r.set_ylim(ylim[0] / SECTOR_ANGLE_DEG, ylim[1] / SECTOR_ANGLE_DEG)

            # Rescale bottom subplot Y (delta angle)
            ax2 = self.ax_motor_sec
            vis_d1 = self._motor_m1_delta[mask]
            vis_d2 = self._motor_m2_delta[mask]
            d_lo = min(vis_d1.min(), vis_d2.min())
            d_hi = max(vis_d1.max(), vis_d2.max())
            d_margin = (d_hi - d_lo) * 0.1 if d_hi > d_lo else 10
            ax2.set_ylim(d_lo - d_margin, d_hi + d_margin)
            # Sync bottom right axis (Δ sectors)
            ylim2 = ax2.get_ylim()
            if hasattr(self, '_motor_ax2_r'):
                self._motor_ax2_r.set_ylim(ylim2[0] / SECTOR_ANGLE_DEG, ylim2[1] / SECTOR_ANGLE_DEG)
        finally:
            self._motor_updating = False

        self.canvas_motor.draw_idle()

    def _on_motor_scroll(self, event):
        """Scroll wheel zooms shared X-axis centered on mouse position.
        Works from either subplot. Y auto-rescales via xlim_changed callback."""
        if event.inaxes is None:
            return
        # Accept scroll on either subplot or their twin axes
        valid_axes = {self.ax_motor_deg, self.ax_motor_sec}
        if hasattr(self, '_motor_ax1_r'):
            valid_axes.add(self._motor_ax1_r)
        if hasattr(self, '_motor_ax2_r'):
            valid_axes.add(self._motor_ax2_r)
        if event.inaxes not in valid_axes:
            return

        # Read current X range from the shared axis (top subplot owns it)
        ax1 = self.ax_motor_deg
        x_lo, x_hi = ax1.get_xlim()
        factor = 0.8 if event.button == "up" else 1.25
        x_mouse = event.xdata
        if x_mouse is None:
            x_mouse = (x_lo + x_hi) / 2.0
        new_lo = x_mouse - (x_mouse - x_lo) * factor
        new_hi = x_mouse + (x_hi - x_mouse) * factor
        # Clamp to full data range
        full_lo, full_hi = self._motor_x_full
        new_lo = max(new_lo, full_lo - (full_hi - full_lo) * 0.02)
        new_hi = min(new_hi, full_hi + (full_hi - full_lo) * 0.02)
        # Setting xlim on the top axis propagates to bottom via sharex
        ax1.set_xlim(new_lo, new_hi)
        # xlim_changed callback handles Y-rescale for both subplots

    def _on_motor_dblclick(self, event):
        """Double-click on either subplot resets zoom to full X range."""
        if not event.dblclick:
            return
        valid_axes = {self.ax_motor_deg, self.ax_motor_sec}
        if hasattr(self, '_motor_ax1_r'):
            valid_axes.add(self._motor_ax1_r)
        if hasattr(self, '_motor_ax2_r'):
            valid_axes.add(self._motor_ax2_r)
        if event.inaxes in valid_axes:
            full_lo, full_hi = self._motor_x_full
            self.ax_motor_deg.set_xlim(full_lo, full_hi)

    def _draw_motor(self):
        if not self.motor: return
        n = len(self.motor["m1_deg"])
        dt = CYCLE_TIME_MS / 1000.0
        ts = [i * dt for i in range(n)]

        # Delta-angle per cycle (angular velocity indicator)
        m1_delta = [0.0] + [self.motor["m1_deg"][i] - self.motor["m1_deg"][i-1] for i in range(1, n)]
        m2_delta = [0.0] + [self.motor["m2_deg"][i] - self.motor["m2_deg"][i-1] for i in range(1, n)]

        self.motor_info_var.set(
            f"M1: {self.motor['m1_deg'][0]:.1f}\u00b0\u2013{max(self.motor['m1_deg']):.1f}\u00b0  "
            f"({self.motor['m1_deg'][0]/360:.1f}\u2013{max(self.motor['m1_deg'])/360:.1f} rev)  |  "
            f"M2: {self.motor['m2_deg'][0]:.1f}\u00b0\u2013{max(self.motor['m2_deg']):.1f}\u00b0  "
            f"({self.motor['m2_deg'][0]/360:.1f}\u2013{max(self.motor['m2_deg'])/360:.1f} rev)  |  "
            f"Sub-sector cycles: {sum(1 for d in m1_delta if abs(d) < 72 and abs(d) > 0.01)}"
            f"+{sum(1 for d in m2_delta if abs(d) < 72 and abs(d) > 0.01)} "
            f"of {n} ({sum(1 for d1, d2 in zip(m1_delta, m2_delta) if (abs(d1) < 72 and abs(d1) > 0.01) or (abs(d2) < 72 and abs(d2) > 0.01))} any)"
        )

        ax1 = self.ax_motor_deg; ax1.clear()
        ax1.plot(ts, self.motor["m1_deg"], "-", color="#2166ac", lw=0.8, label="M1 motor angle [\u00b0]")
        ax1.plot(ts, self.motor["m2_deg"], "-", color="#d85a30", lw=0.8, label="M2 motor angle [\u00b0]")

        # Store data arrays for the auto-rescale callback
        self._motor_ts = np.array(ts)
        self._motor_m1_deg = np.array(self.motor["m1_deg"])
        self._motor_m2_deg = np.array(self.motor["m2_deg"])
        self._motor_m1_delta = np.array(m1_delta)
        self._motor_m2_delta = np.array(m2_delta)

        # Initial Y-axis setup
        self._apply_motor_yticks(ax1, 0, max(self.motor["m1_deg"] + self.motor["m2_deg"]) * 1.05)

        ax1.set_ylabel("Motor angle [\u00b0]")
        # Right axis: commutation sectors
        ax1_r = ax1.twinx()
        ax1_r.set_ylim(ax1.get_ylim()[0] / SECTOR_ANGLE_DEG, ax1.get_ylim()[1] / SECTOR_ANGLE_DEG)
        ax1_r.set_ylabel("Commutation sectors (72\u00b0 each)")
        self._motor_ax1_r = ax1_r
        ax1.set_title("Motor Angular Position (major grid = labeled revolutions, minor grid = 1 rev = 360\u00b0)")
        ax1.legend(fontsize=8, loc="upper left")
        # Hide X tick labels on top subplot — bottom subplot shows the shared time axis
        plt.setp(ax1.get_xticklabels(), visible=False)

        # Connect xlim_changed to auto-rescale Y for both subplots
        if hasattr(self, "_motor_xlim_cid"):
            ax1.callbacks.disconnect(self._motor_xlim_cid)
        self._motor_xlim_cid = ax1.callbacks.connect("xlim_changed", self._on_motor_xlim_changed)

        ax2 = self.ax_motor_sec; ax2.clear()
        ax2.plot(ts, m1_delta, "-", color="#2166ac", lw=0.6, label="M1 \u0394angle/cycle [\u00b0]")
        ax2.plot(ts, m2_delta, "-", color="#d85a30", lw=0.6, label="M2 \u0394angle/cycle [\u00b0]")
        ax2.axhline(72.0, color="#e24b4a", lw=0.8, ls="--", alpha=0.4, label="1 commutation sector (72\u00b0)")
        ax2.axhline(-72.0, color="#e24b4a", lw=0.8, ls="--", alpha=0.4)

        # Highlight sub-sector motion regions: where |Δangle| < 72° for BOTH motors
        # These are the critical zones where commutator granularity dominates
        m1_abs = np.abs(self._motor_m1_delta)
        m2_abs = np.abs(self._motor_m2_delta)
        sub_sector = (m1_abs < 72.0) | (m2_abs < 72.0)
        # Find contiguous regions
        changes = np.diff(sub_sector.astype(int))
        starts = np.where(changes == 1)[0] + 1
        ends = np.where(changes == -1)[0] + 1
        if sub_sector[0]:
            starts = np.concatenate(([0], starts))
        if sub_sector[-1]:
            ends = np.concatenate((ends, [n - 1]))
        for s_i, e_i in zip(starts, ends):
            ax2.axvspan(ts[s_i], ts[min(e_i, n - 1)],
                        alpha=0.08, color="#e24b4a", zorder=0)
        # Add a single legend entry for the highlight
        if len(starts) > 0:
            ax2.fill_between([], [], [], alpha=0.15, color="#e24b4a",
                             label="Sub-sector motion (\u0394 < 72\u00b0)")

        ax2.set_xlabel("Time [s]"); ax2.set_ylabel("\u0394angle / cycle [\u00b0]")
        # Right Y-axis: show same data in commutation sectors
        ax2_r = ax2.twinx()
        ylim2 = ax2.get_ylim()
        ax2_r.set_ylim(ylim2[0] / SECTOR_ANGLE_DEG, ylim2[1] / SECTOR_ANGLE_DEG)
        ax2_r.set_ylabel("\u0394 commutation sectors / cycle")
        self._motor_ax2_r = ax2_r
        ax2.set_title("Motor Angular Change per 100 Hz Cycle "
                       "(red band = sub-sector motion, commutation-critical)")
        ax2.legend(fontsize=8, loc="upper left"); ax2.grid(True, alpha=0.3)

        # Store full X range for reset
        self._motor_x_full = (ts[0], ts[-1])

        # Connect mouse scroll zoom and double-click reset
        if hasattr(self, "_motor_scroll_cid"):
            self.canvas_motor.mpl_disconnect(self._motor_scroll_cid)
        if hasattr(self, "_motor_dblclick_cid"):
            self.canvas_motor.mpl_disconnect(self._motor_dblclick_cid)
        self._motor_scroll_cid = self.canvas_motor.mpl_connect("scroll_event", self._on_motor_scroll)
        self._motor_dblclick_cid = self.canvas_motor.mpl_connect("button_press_event", self._on_motor_dblclick)

        self.fig_motor.tight_layout(); self.canvas_motor.draw()

    def _fill_table(self):
        for item in self.tree.get_children(): self.tree.delete(item)
        for r in self.results:
            self.tree.insert("", tk.END, values=(
                r["step"],r["point"],r["cmd"],r["vel"],r["blending"],
                f'{r["X"]:.1f}',f'{r["Y"]:.1f}',f'{r["Z"]:.1f}',
                f'{r["dist_to_tracker"]:.1f}',f'{r["segment_dist"]:.1f}',f'{r["time"]:.3f}'))

    # --- export -------------------------------------------------------------
    def _export_csv(self):
        if not self.results: messagebox.showinfo("Info","Load files first."); return
        path=filedialog.asksaveasfilename(defaultextension=".csv",filetypes=[("CSV","*.csv")],title="Export CSV")
        if not path: return
        with open(path,"w",newline="") as f:
            w=csv.writer(f,delimiter=";")
            w.writerow(["Step","Point","Cmd","Velocity","Blending","X_mm","Y_mm","Z_mm","DistTracker_mm","SegDist_mm","Time_s"])
            for r in self.results:
                w.writerow([r["step"],r["point"],r["cmd"],r["vel"],r["blending"],
                    f'{r["X"]:.3f}',f'{r["Y"]:.3f}',f'{r["Z"]:.3f}',
                    f'{r["dist_to_tracker"]:.3f}',f'{r["segment_dist"]:.3f}',f'{r["time"]:.6f}'])
        self.status_var.set(f"Exported CSV to {path}")

    def _export_mv_profile(self):
        if not self.results: messagebox.showinfo("Info","Load files first."); return
        path=filedialog.asksaveasfilename(defaultextension=".csv",filetypes=[("CSV","*.csv")],
            title="Export MV profile",initialfile="mv_distance_profile.csv")
        if not path: return
        with open(path,"w",newline="") as f:
            w=csv.writer(f,delimiter=";")
            w.writerow(["Time_s","Distance_mm","Distance_mm_clamped"])
            for r in self.results:
                w.writerow([f'{r["time"]:.6f}',f'{r["dist_to_tracker"]:.3f}',f'{clamp_mv(r["dist_to_tracker"]):.3f}'])
        self.status_var.set(f"Exported MV profile to {path}")

    def _export_cpp_lut(self):
        if not self.lut: messagebox.showinfo("Info","Load files first."); return
        tp=(self.tracker_x.get(),self.tracker_y.get(),self.tracker_z.get())
        header_text=generate_cpp_header(self.lut, self.src_name, tp)
        path=filedialog.asksaveasfilename(defaultextension=".h",
            filetypes=[("C/C++ Header","*.h"),("All","*.*")],
            title="Export C++ Profile Replay LUT",initialfile="VarioProfileReplayLUT.h")
        if not path: return
        with open(path,"w") as f: f.write(header_text)
        n=len(self.lut)
        self.status_var.set(
            f"Exported C++ LUT to {os.path.basename(path)} \u2014 "
            f"{n} entries, {n*4} bytes, {n*CYCLE_TIME_MS/1000.0:.2f} s")

    def _export_motor_csv(self):
        if not self.motor:
            messagebox.showinfo("Info", "No motor profiles computed (load files and zoom tables first).")
            return
        path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV","*.csv")],
            title="Export Motor Profiles", initialfile="motor_profiles.csv")
        if not path:
            return
        n = len(self.motor["m1_incr"])
        dt = CYCLE_TIME_MS / 1000.0
        with open(path, "w", newline="") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow([
                "Time_s", "Distance_mm",
                "M1_incr", "M1_linear_mm", "M1_angle_deg", "M1_sectors",
                "M1_delta_deg", "M1_delta_sectors",
                "M2_incr", "M2_linear_mm", "M2_angle_deg", "M2_sectors",
                "M2_delta_deg", "M2_delta_sectors",
            ])
            for i in range(n):
                m1_dd = self.motor['m1_deg'][i] - self.motor['m1_deg'][i-1] if i > 0 else 0.0
                m2_dd = self.motor['m2_deg'][i] - self.motor['m2_deg'][i-1] if i > 0 else 0.0
                w.writerow([
                    f"{i * dt:.6f}", f"{self.lut[i]}",
                    f"{self.motor['m1_incr'][i]:.1f}",
                    f"{self.motor['m1_mm'][i]:.4f}",
                    f"{self.motor['m1_deg'][i]:.2f}",
                    f"{self.motor['m1_sectors'][i]:.3f}",
                    f"{m1_dd:.2f}",
                    f"{m1_dd / SECTOR_ANGLE_DEG:.4f}",
                    f"{self.motor['m2_incr'][i]:.1f}",
                    f"{self.motor['m2_mm'][i]:.4f}",
                    f"{self.motor['m2_deg'][i]:.2f}",
                    f"{self.motor['m2_sectors'][i]:.3f}",
                    f"{m2_dd:.2f}",
                    f"{m2_dd / SECTOR_ANGLE_DEG:.4f}",
                ])
        self.status_var.set(f"Exported motor profiles to {path}  ({n} entries)")

    def _export_motor_xlsx(self):
        """Export motor movement profile as a professionally formatted Excel workbook
        for sharing with the motor supplier (Faulhaber)."""
        if not self.motor:
            messagebox.showinfo("Info", "No motor profiles computed (load files and zoom tables first).")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
            from openpyxl.chart import ScatterChart, Reference, Series
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("Error", "openpyxl is required for Excel export.\n"
                                 "Install with:  pip install openpyxl")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("All", "*.*")],
            title="Export Motor Profile for Supplier",
            initialfile="MiniVario_Motor_Movement_Profile.xlsx")
        if not path:
            return

        n = len(self.motor["m1_incr"])
        dt = CYCLE_TIME_MS / 1000.0

        # Precompute deltas
        m1_delta = [0.0] + [self.motor['m1_deg'][i] - self.motor['m1_deg'][i-1] for i in range(1, n)]
        m2_delta = [0.0] + [self.motor['m2_deg'][i] - self.motor['m2_deg'][i-1] for i in range(1, n)]

        wb = Workbook()

        # ── Styles ──
        hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill("solid", fgColor="2F5496")
        data_font = Font(name="Arial", size=9)
        info_font = Font(name="Arial", size=10)
        info_bold = Font(name="Arial", size=10, bold=True)
        title_font = Font(name="Arial", size=14, bold=True, color="2F5496")
        subtitle_font = Font(name="Arial", size=10, italic=True, color="666666")
        thin = Side(style='thin', color='C0C0C0')
        border = Border(top=thin, bottom=thin, left=thin, right=thin)
        warn_fill = PatternFill("solid", fgColor="FCE4EC")
        ok_fill = PatternFill("solid", fgColor="E8F5E9")

        # ====================================================================
        #  Sheet 1: Description
        # ====================================================================
        ws_info = wb.active
        ws_info.title = "Description"
        ws_info.column_dimensions['A'].width = 38
        ws_info.column_dimensions['B'].width = 65

        ws_info['A1'] = "MiniVario Motor Movement Profile"
        ws_info['A1'].font = title_font
        ws_info.merge_cells('A1:B1')

        ws_info['A2'] = "Representative target profile from customer automation application"
        ws_info['A2'].font = subtitle_font
        ws_info.merge_cells('A2:B2')

        desc_rows = [
            ("", ""),
            ("IMPORTANT NOTE", ""),
            ("", "This is NOT a measured motor movement recording. It is a computed"),
            ("", "target profile derived from a KUKA robot program as follows:"),
            ("", ""),
            ("1. Source", "KUKA KRL robot program (.src/.dat) defining the TCP path"),
            ("2. Distance computation", "Euclidean distance from Laser Tracker (AT960) to robot TCP"),
            ("3. Zoom table lookup", "Distance \u2192 lens carriage target position via zoom curves (S8/S23)"),
            ("4. Kinematic conversion", "Lens position \u2192 motor angular position via gear + pulley chain"),
            ("", ""),
            ("", "The profile represents the ideal, commanded motor trajectory that the"),
            ("", "MiniVario drives must follow during this specific automation application."),
            ("", "It shows the motion patterns, speeds, direction reversals, and the"),
            ("", "relationship between carriage micro-motion and commutation sectors."),
            ("", ""),
            ("SYSTEM PARAMETERS", ""),
            ("Motor", "Faulhaber 1016-012-SR"),
            ("Commutator", f"{COMMUTATOR_SEGMENTS} segments, {SECTOR_ANGLE_DEG:.0f}\u00b0 per sector"),
            ("Gear ratio", f"{GEAR_RATIO:.0f}:1 (planetary gearbox)"),
            ("Pulley diameter", f"{PULLEY_DIAMETER_MM:.0f} mm (circumference {math.pi * PULLEY_DIAMETER_MM:.2f} mm)"),
            ("Motor rev / mm linear", f"{MOTOR_REV_PER_MM:.4f}"),
            ("Linear per comm. sector", f"{LINEAR_PER_SECTOR_UM:.1f} \u00b5m"),
            ("Encoder resolution", "0.5 \u00b5m / increment"),
            ("Control loop rate", "200 Hz (servo), 100 Hz (distance generator)"),
            ("", ""),
            ("PROFILE SUMMARY", ""),
            ("Source program", self.src_name or "(unknown)"),
            ("Duration", f"{n * dt:.2f} s ({n} cycles @ {GENERATOR_RATE_HZ} Hz)"),
            ("M1 (Zoom1/S8) range",
             f"{min(self.motor['m1_deg']):.0f}\u00b0 \u2013 {max(self.motor['m1_deg']):.0f}\u00b0  "
             f"({min(self.motor['m1_deg'])/360:.1f} \u2013 {max(self.motor['m1_deg'])/360:.1f} rev)"),
            ("M2 (Zoom2/S23) range",
             f"{min(self.motor['m2_deg']):.0f}\u00b0 \u2013 {max(self.motor['m2_deg']):.0f}\u00b0  "
             f"({min(self.motor['m2_deg'])/360:.1f} \u2013 {max(self.motor['m2_deg'])/360:.1f} rev)"),
            ("M1 linear travel range",
             f"{min(self.motor['m1_mm']):.3f} \u2013 {max(self.motor['m1_mm']):.3f} mm"),
            ("M2 linear travel range",
             f"{min(self.motor['m2_mm']):.3f} \u2013 {max(self.motor['m2_mm']):.3f} mm"),
            ("Sub-sector cycles (M1)",
             f"{sum(1 for d in m1_delta if 0.01 < abs(d) < 72)} of {n}"),
            ("Sub-sector cycles (M2)",
             f"{sum(1 for d in m2_delta if 0.01 < abs(d) < 72)} of {n}"),
        ]
        for row_idx, (label, value) in enumerate(desc_rows, start=4):
            c_a = ws_info.cell(row=row_idx, column=1, value=label)
            c_b = ws_info.cell(row=row_idx, column=2, value=value)
            if label in ("IMPORTANT NOTE", "SYSTEM PARAMETERS", "PROFILE SUMMARY"):
                c_a.font = info_bold
            else:
                c_a.font = info_bold if label else info_font
                c_b.font = info_font

        # ====================================================================
        #  Sheet 2: Motor 1 (Zoom1 / S8) profile
        # ====================================================================
        ws_m1 = wb.create_sheet("M1 – Zoom1 (S8)")
        headers_m = [
            ("Time [s]", 10),
            ("Focus dist [mm]", 14),
            ("Lens pos [incr]", 14),
            ("Linear [mm]", 12),
            ("Motor angle [\u00b0]", 14),
            ("Motor rev", 10),
            ("Comm. sector", 12),
            ("\u0394angle/cycle [\u00b0]", 15),
            ("\u0394sectors/cycle", 14),
            ("Sub-sector?", 11),
        ]
        for c, (h, w) in enumerate(headers_m, 1):
            cell = ws_m1.cell(row=1, column=c, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            ws_m1.column_dimensions[get_column_letter(c)].width = w

        for i in range(n):
            dd = m1_delta[i]
            ds = dd / SECTOR_ANGLE_DEG
            sub = "YES" if (0.01 < abs(dd) < SECTOR_ANGLE_DEG) else ""
            row = i + 2
            vals = [
                round(i * dt, 4), self.lut[i],
                round(self.motor['m1_incr'][i], 1), round(self.motor['m1_mm'][i], 4),
                round(self.motor['m1_deg'][i], 2), round(self.motor['m1_deg'][i] / 360.0, 3),
                round(self.motor['m1_sectors'][i], 3),
                round(dd, 2), round(ds, 4), sub
            ]
            for c, v in enumerate(vals, 1):
                cell = ws_m1.cell(row=row, column=c, value=v)
                cell.font = data_font; cell.border = border
                if c >= 4:
                    cell.alignment = Alignment(horizontal='right')
                if sub == "YES" and c == 10:
                    cell.fill = warn_fill
                    cell.font = Font(name="Arial", size=9, color="C62828")

        ws_m1.auto_filter.ref = f"A1:J{n+1}"
        ws_m1.freeze_panes = "A2"

        # ====================================================================
        #  Sheet 3: Motor 2 (Zoom2 / S23) profile
        # ====================================================================
        ws_m2 = wb.create_sheet("M2 – Zoom2 (S23)")
        for c, (h, w) in enumerate(headers_m, 1):
            cell = ws_m2.cell(row=1, column=c, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            ws_m2.column_dimensions[get_column_letter(c)].width = w

        for i in range(n):
            dd = m2_delta[i]
            ds = dd / SECTOR_ANGLE_DEG
            sub = "YES" if (0.01 < abs(dd) < SECTOR_ANGLE_DEG) else ""
            row = i + 2
            vals = [
                round(i * dt, 4), self.lut[i],
                round(self.motor['m2_incr'][i], 1), round(self.motor['m2_mm'][i], 4),
                round(self.motor['m2_deg'][i], 2), round(self.motor['m2_deg'][i] / 360.0, 3),
                round(self.motor['m2_sectors'][i], 3),
                round(dd, 2), round(ds, 4), sub
            ]
            for c, v in enumerate(vals, 1):
                cell = ws_m2.cell(row=row, column=c, value=v)
                cell.font = data_font; cell.border = border
                if c >= 4:
                    cell.alignment = Alignment(horizontal='right')
                if sub == "YES" and c == 10:
                    cell.fill = warn_fill
                    cell.font = Font(name="Arial", size=9, color="C62828")

        ws_m2.auto_filter.ref = f"A1:J{n+1}"
        ws_m2.freeze_panes = "A2"

        # ====================================================================
        #  Sheet 4: Charts
        # ====================================================================
        ws_chart = wb.create_sheet("Charts")

        # Write compact chart data (time, M1 angle, M2 angle, M1 delta, M2 delta)
        ch_headers = ["Time [s]", "M1 angle [\u00b0]", "M2 angle [\u00b0]",
                      "M1 \u0394angle [\u00b0/cycle]", "M2 \u0394angle [\u00b0/cycle]"]
        for c, h in enumerate(ch_headers, 1):
            cell = ws_chart.cell(row=1, column=c, value=h)
            cell.font = Font(name="Arial", bold=True, size=9)
        # Downsample for chart performance (max ~2000 points)
        step = max(1, n // 2000)
        chart_n = 0
        for i in range(0, n, step):
            chart_n += 1
            row = chart_n + 1
            ws_chart.cell(row=row, column=1, value=round(i * dt, 4))
            ws_chart.cell(row=row, column=2, value=round(self.motor['m1_deg'][i], 1))
            ws_chart.cell(row=row, column=3, value=round(self.motor['m2_deg'][i], 1))
            ws_chart.cell(row=row, column=4, value=round(m1_delta[i], 1))
            ws_chart.cell(row=row, column=5, value=round(m2_delta[i], 1))

        # Chart 1: Motor angular position
        chart1 = ScatterChart()
        chart1.title = "Motor Angular Position (Target Profile)"
        chart1.x_axis.title = "Time [s]"
        chart1.y_axis.title = "Motor angle [\u00b0]"
        chart1.width = 28; chart1.height = 14
        chart1.style = 2
        xv = Reference(ws_chart, min_col=1, min_row=2, max_row=chart_n + 1)
        s1 = Series(Reference(ws_chart, min_col=2, min_row=2, max_row=chart_n + 1), xv, title="M1 (Zoom1/S8)")
        s1.graphicalProperties.line.width = 12000
        s2 = Series(Reference(ws_chart, min_col=3, min_row=2, max_row=chart_n + 1), xv, title="M2 (Zoom2/S23)")
        s2.graphicalProperties.line.width = 12000
        chart1.series.append(s1)
        chart1.series.append(s2)
        ws_chart.add_chart(chart1, "G1")

        # Chart 2: Delta angle per cycle
        chart2 = ScatterChart()
        chart2.title = "Motor Angular Change per Cycle (\u0394angle, 72\u00b0 = 1 comm. sector)"
        chart2.x_axis.title = "Time [s]"
        chart2.y_axis.title = "\u0394angle / cycle [\u00b0]"
        chart2.width = 28; chart2.height = 14
        chart2.style = 2
        xv2 = Reference(ws_chart, min_col=1, min_row=2, max_row=chart_n + 1)
        s3 = Series(Reference(ws_chart, min_col=4, min_row=2, max_row=chart_n + 1), xv2, title="M1 \u0394angle")
        s3.graphicalProperties.line.width = 10000
        s4 = Series(Reference(ws_chart, min_col=5, min_row=2, max_row=chart_n + 1), xv2, title="M2 \u0394angle")
        s4.graphicalProperties.line.width = 10000
        chart2.series.append(s3)
        chart2.series.append(s4)
        ws_chart.add_chart(chart2, "G18")

        # Hide chart data columns
        for c in range(1, 6):
            ws_chart.column_dimensions[get_column_letter(c)].width = 0.5

        # ── Save ──
        wb.save(path)
        self.status_var.set(
            f"Exported Excel motor profile to {os.path.basename(path)} \u2014 "
            f"{n} entries, M1+M2, {n * dt:.1f} s"
        )


if __name__ == "__main__":
    app = App()
    app.mainloop()
